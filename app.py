"""
Excel Agent Online - Web版本Excel处理Agent
- 交互界面：网页端
- API配置：支持配置API Key连接外部模型
- 技能定制：通过输入文字让AI生成Excel处理技能
- 文件处理：处理完的文件提供下载链接
- 图片显示：生成的图片直接在网页端展示
"""
import os
import sys
import json
import time
import uuid
import shutil
import webbrowser
import importlib.util
from pathlib import Path
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS

# 添加项目根目录到路径
# PyInstaller打包兼容：frozen时使用exe所在目录，否则使用脚本所在目录
if getattr(sys, 'frozen', False):
    # 打包后的exe运行时，_MEIPASS是临时解压目录（只读资源），exe_dir是exe所在目录（用户数据）
    MEIPASS = Path(sys._MEIPASS).resolve()
    EXE_DIR = Path(sys.executable).parent.resolve()
    # 只读资源（静态文件、内置技能、AI技能规范）从临时解压目录读取
    STATIC_SRC_DIR = MEIPASS
    # 可写目录（用户上传、输出、待处理、AI生成技能）放在exe同级目录
    BASE_DIR = EXE_DIR
else:
    BASE_DIR = Path(__file__).parent.resolve()
    STATIC_SRC_DIR = BASE_DIR

SKILL_DIR = BASE_DIR / 'skill'
PENDING_DIR = BASE_DIR / 'pending'
OUTPUT_DIR = BASE_DIR / 'output'
UPLOAD_DIR = BASE_DIR / 'uploads'
AI_SKILL_DIR = BASE_DIR / 'ai_skills'

for d in [SKILL_DIR, PENDING_DIR, OUTPUT_DIR, UPLOAD_DIR, AI_SKILL_DIR]:
    d.mkdir(exist_ok=True)

# 打包模式：首次运行时，将打包的内置技能复制到exe所在目录
if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
    _bundle = Path(sys._MEIPASS).resolve()
    _bundle_skill = _bundle / 'skill'
    _bundle_ai = _bundle / 'ai_skills'
    if _bundle_skill.exists():
        for _item in _bundle_skill.iterdir():
            if _item.is_dir() and not (SKILL_DIR / _item.name).exists():
                shutil.copytree(str(_item), str(SKILL_DIR / _item.name))
    if _bundle_ai.exists():
        for _item in _bundle_ai.iterdir():
            if _item.is_dir() and not _item.name.startswith('_') and not (AI_SKILL_DIR / _item.name).exists():
                shutil.copytree(str(_item), str(AI_SKILL_DIR / _item.name))
            elif _item.is_file() and not (AI_SKILL_DIR / _item.name).exists():
                shutil.copy2(str(_item), str(AI_SKILL_DIR / _item.name))
    # 复制使用说明.txt到exe所在目录
    _readme_src = _bundle / '使用说明.txt'
    if _readme_src.exists() and not (EXE_DIR / '使用说明.txt').exists():
        shutil.copy2(str(_readme_src), str(EXE_DIR / '使用说明.txt'))

app = Flask(__name__, static_folder=str(STATIC_SRC_DIR / 'static'))
CORS(app)

# 全局状态
api_config = {
    'provider': '',       # 'openai' / 'anthropic' / 'deepseek'
    'api_key': '',
    'base_url': '',       # 自定义API地址
    'model': '',          # 使用的模型名称
}

# 配置持久化文件
CONFIG_FILE = BASE_DIR / 'config.json'

def _load_config_from_file():
    """启动时从文件加载API配置"""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                saved = json.load(f)
            for key in ['provider', 'api_key', 'base_url', 'model']:
                if key in saved and saved[key]:
                    api_config[key] = saved[key]
            print(f"  已从配置文件加载API配置 (提供商: {api_config['provider']})")
        except Exception as e:
            print(f"  加载配置文件失败: {e}")

def _save_config_to_file():
    """将API配置持久化到文件"""
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(api_config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  保存配置文件失败: {e}")

def _clear_config_file():
    """清除配置文件"""
    try:
        if CONFIG_FILE.exists():
            CONFIG_FILE.unlink()
    except Exception:
        pass

# 启动时加载配置
_load_config_from_file()

def detect_time_column(headers):
    """检测时间列"""
    for i, h in enumerate(headers):
        if h and '时间' in str(h):
            return i
    return -1

def detect_numeric_columns(all_data, headers, time_col):
    """检测数值列（排除编号、文本类型列）"""
    value_cols = []
    skip_keywords = ['编号', '序号', '类型', '型号', '备注', '名称']
    for i, h in enumerate(headers):
        if i == time_col:
            continue
        h_str = str(h).strip()
        if not h_str:
            continue
        skip = False
        for kw in skip_keywords:
            if kw in h_str:
                skip = True
                break
        if skip:
            continue
        # 检查是否有70%以上的数值
        numeric_count = 0
        total_count = 0
        for r in range(1, min(len(all_data), 20)):
            if i >= len(all_data[r]):
                continue
            val = all_data[r][i]
            if val is None or (isinstance(val, str) and val.strip() in ['', '--', '-']):
                continue
            total_count += 1
            if isinstance(val, (int, float)):
                numeric_count += 1
            elif isinstance(val, str):
                try:
                    float(val.strip().replace('℃', '').replace('°C', '').strip())
                    numeric_count += 1
                except ValueError:
                    pass
        if total_count > 0 and numeric_count / total_count >= 0.5:
            value_cols.append(i)
        elif any(kw in h_str for kw in ['℃', '°C', '温度', 'T1', 'T2', 'T3', 'T4', 'T5']):
            value_cols.append(i)
    return value_cols

def process_wensheng(input_path, output_path):
    """温升技能处理函数"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import LineChart, Reference
    from datetime import datetime
    import xlrd

    ext = os.path.splitext(input_path)[1].lower()
    all_data = []
    sheet_name = 'Sheet1'

    if ext == '.xls':
        wb_in = xlrd.open_workbook(input_path)
        sheet_in = wb_in.sheet_by_index(0)
        sheet_name = sheet_in.name
        for r in range(sheet_in.nrows):
            row = [sheet_in.cell(r, c).value for c in range(sheet_in.ncols)]
            all_data.append(row)
    else:
        wb_in = load_workbook(input_path, data_only=True)
        ws_in = wb_in.active
        sheet_name = ws_in.title
        for row in ws_in.iter_rows(values_only=True):
            all_data.append(list(row))

    if not all_data:
        return None, []

    headers = [str(h).strip() if h is not None else '' for h in all_data[0]]
    time_col = detect_time_column(headers)
    value_cols = detect_numeric_columns(all_data, headers, time_col)
    nrows = len(all_data)
    ncols = len(all_data[0]) if nrows > 0 else 0

    if time_col < 0 or not value_cols:
        return None, []

    # 创建输出工作簿
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = sheet_name
    for c, h in enumerate(headers):
        ws_out.cell(row=1, column=c + 1, value=all_data[0][c])

    first_time_str = None
    for r in range(1, nrows):
        for c in range(ncols):
            if c >= len(all_data[r]):
                continue
            val = all_data[r][c]
            if c == time_col:
                dt_str = str(val).strip() if val is not None else ''
                if not dt_str or dt_str in ['--', '-']:
                    ws_out.cell(row=r+1, column=c+1, value=val)
                    continue
                # 解析时间
                parsed_dt = None
                for fmt in ["%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                           "%H:%M:%S.%f", "%H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
                    try:
                        parsed_dt = datetime.strptime(dt_str, fmt)
                        break
                    except ValueError:
                        continue
                if parsed_dt is None:
                    ws_out.cell(row=r+1, column=c+1, value=val)
                    continue
                if r == 1:
                    first_time_str = dt_str
                    if parsed_dt.year == 1900 and parsed_dt.month == 1 and parsed_dt.day == 1:
                        ws_out.cell(row=r+1, column=c+1, value=dt_str)
                    else:
                        ws_out.cell(row=r+1, column=c+1,
                                   value=f"{parsed_dt.year}/{parsed_dt.month}/{parsed_dt.day} "
                                         f"{parsed_dt.hour:02d}:{parsed_dt.minute:02d}:{parsed_dt.second:02d}")
                else:
                    if parsed_dt.microsecond > 0:
                        time_only = f"{parsed_dt.hour:02d}:{parsed_dt.minute:02d}:{parsed_dt.second:02d}.{parsed_dt.microsecond//1000:03d}"
                    else:
                        time_only = f"{parsed_dt.hour:02d}:{parsed_dt.minute:02d}:{parsed_dt.second:02d}"
                    ws_out.cell(row=r+1, column=c+1, value=time_only)
            elif c in value_cols:
                converted = val
                if isinstance(val, str):
                    try:
                        converted = float(val.strip())
                    except (ValueError, AttributeError):
                        converted = val
                ws_out.cell(row=r+1, column=c+1, value=converted)
            else:
                ws_out.cell(row=r+1, column=c+1, value=val)

    # 插入折线图
    chart = LineChart()
    chart.title = "温升温度曲线"
    chart.y_axis.title = "温度"
    chart.x_axis.title = "时间"
    chart.style = 10
    chart.width = 25
    chart.height = 15
    x_data = Reference(ws_out, min_col=time_col+1, min_row=2, max_row=nrows)
    for vc in value_cols:
        data = Reference(ws_out, min_col=vc+1, min_row=1, max_row=nrows)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(x_data)
    ws_out.add_chart(chart, f"A{nrows+3}")

    wb_out.save(output_path)
    return output_path, []


def try_generate_skill_with_ai(skill_name, description, example_before_data='', example_after_data=''):
    """使用AI生成技能代码，可选传入示例文件数据供参考"""
    if not api_config['api_key'] or not api_config['provider']:
        return None, "未配置API Key，请先配置"

    skill_code = None
    error = None

    system_prompt = """你是一个Python代码生成专家。用户会给你一个Excel/文本处理需求，你需要生成一个Python函数。

要求：
1. 函数名必须是 process(input_path, output_path)，接收输入文件路径和输出文件路径
2. 使用 openpyxl 处理 .xlsx 文件，使用 xlrd 处理 .xls 文件
3. 必须支持同时处理 .xls、.xlsx 和 .txt 三种格式
4. 当输入为 .txt 文件时，需要读取文本文件内容（自动检测分隔符如制表符、逗号、空格、竖线等），解析为表格数据，然后转换为 .xlsx 格式输出
5. 输出必须是 .xlsx 格式
6. 函数返回 (output_path, image_paths_list)
7. 如果需要生成图片，使用 matplotlib，将图片保存到与output_path相同的目录，文件名用 _chart_N.png
8. 不要在代码中使用 print() 函数，改用 logging 模块记录日志
9. 只返回纯Python代码，不要包含任何markdown标记、注释或解释
10. import 语句全部放在函数内部
11. 图片保存路径用 os.path.join(os.path.dirname(output_path), f'chart_{idx+1}.png')
12. 处理txt文件时，尝试多种编码读取（utf-8, gbk, gb2312, latin-1等），自动识别表头行

示例格式：
import os
def process(input_path, output_path):
    import openpyxl
    from openpyxl import Workbook, load_workbook
    # ... 处理逻辑
    return output_path, []
"""

    user_prompt = f"""请为以下Excel处理需求生成Python处理函数：

技能名称：{skill_name}
需求描述：{description}"""
    if example_before_data:
        user_prompt += f"""

处理前的示例文件数据（表头+前3行）：
{example_before_data}"""
    if example_after_data:
        user_prompt += f"""

处理后的示例文件数据（表头+前3行，供参考输出格式）：
{example_after_data}"""

    user_prompt += "\n\n请直接返回可执行的Python函数代码，不要任何解释。"

    try:
        import requests as _requests
        provider, base_url, model = _get_provider_config()

        if _is_anthropic():
            resp = _requests.post(
                f"{base_url}/v1/messages",
                headers={
                    "x-api-key": api_config['api_key'],
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json"
                },
                json={
                    "model": model,
                    "system": system_prompt,
                    "messages": [{"role": "user", "content": user_prompt}],
                    "max_tokens": 4096,
                    "temperature": 0.1
                },
                timeout=60
            )
            resp.raise_for_status()
            skill_code = resp.json()['content'][0]['text']
        else:
            # 所有OpenAI兼容的提供商（deepseek/openai/qwen/zhipu/moonshot/baichuan/minimax/spark/ollama/custom）
            resp = _requests.post(
                f"{base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_config['api_key']}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": model,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    "temperature": 0.1,
                    "max_tokens": 4096
                },
                timeout=60
            )
            resp.raise_for_status()
            skill_code = resp.json()['choices'][0]['message']['content']
    except Exception as e:
        return None, f"AI调用失败: {str(e)}"

    if not skill_code:
        return None, "AI未返回代码"

    # 清理代码块标记
    skill_code = skill_code.strip()
    if skill_code.startswith('```'):
        lines = skill_code.split('\n')
        start = 1 if lines[0].startswith('```') else 0
        end = len(lines) - 1 if lines[-1].strip() == '```' else len(lines)
        skill_code = '\n'.join(lines[start:end])

    # 验证代码
    try:
        compile(skill_code, '<string>', 'exec')
    except SyntaxError as e:
        return None, f"生成的代码有语法错误: {e}"

    # 保存技能文件
    skill_path = AI_SKILL_DIR / skill_name
    skill_path.mkdir(exist_ok=True)
    with open(skill_path / 'handler.py', 'w', encoding='utf-8') as f:
        f.write(skill_code)
    with open(skill_path / 'skill.txt', 'w', encoding='utf-8') as f:
        f.write(f"技能名称: {skill_name}\n生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n描述: {description}\n生成方式: AI自动生成")

    return str(skill_path), None


# ===== API 路由 =====

@app.route('/')
def index():
    static_dir = STATIC_SRC_DIR / 'static'
    if not static_dir.exists():
        static_dir.mkdir(exist_ok=True)
    return send_from_directory(str(static_dir), 'index.html')

@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory(str(STATIC_SRC_DIR / 'static'), filename)

@app.route('/api/config', methods=['GET', 'POST'])
def handle_config():
    if request.method == 'GET':
        safe_config = api_config.copy()
        if safe_config.get('api_key'):
            key = safe_config['api_key']
            safe_config['api_key_masked'] = key[:8] + '***' + key[-4:] if len(key) > 12 else '***'
        return jsonify(safe_config)
    try:
        data = request.json
        if data is None:
            return jsonify({'status': 'error', 'message': '请求数据格式错误'}), 400
        if 'provider' in data:
            api_config['provider'] = data['provider']
        if 'api_key' in data and data['api_key']:
            api_config['api_key'] = data['api_key']
        if 'base_url' in data:
            api_config['base_url'] = data['base_url']
        if 'model' in data:
            api_config['model'] = data['model']
        _save_config_to_file()
        return jsonify({'status': 'ok', 'message': '配置已保存'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'保存失败: {str(e)}'}), 500


@app.route('/api/config/clear', methods=['POST'])
def clear_config():
    """清除API配置"""
    api_config['provider'] = ''
    api_config['api_key'] = ''
    api_config['base_url'] = ''
    api_config['model'] = ''
    _clear_config_file()
    return jsonify({'status': 'ok'})



# 提供商默认配置
PROVIDER_DEFAULTS = {
    'deepseek': {'base_url': 'https://api.deepseek.com/v1', 'model': 'deepseek-chat'},
    'openai':   {'base_url': 'https://api.openai.com/v1',   'model': 'gpt-4o'},
    'anthropic': {'base_url': 'https://api.anthropic.com',   'model': 'claude-sonnet-4-20250514'},
    'qwen':     {'base_url': 'https://dashscope.aliyuncs.com/compatible-mode/v1', 'model': 'qwen-plus'},
    'zhipu':    {'base_url': 'https://open.bigmodel.cn/api/paas/v4', 'model': 'glm-4'},
    'moonshot': {'base_url': 'https://api.moonshot.cn/v1',   'model': 'moonshot-v1-8k'},
    'baichuan': {'base_url': 'https://api.baichuan-ai.com/v1', 'model': 'Baichuan4'},
    'minimax':  {'base_url': 'https://api.minimax.chat/v1',  'model': 'MiniMax-Text-01'},
    'spark':    {'base_url': 'https://spark-api-open.xf-yun.com/v1', 'model': 'generalv3.5'},
    'ollama':   {'base_url': 'http://localhost:11434/v1',     'model': 'qwen2.5:7b'},
}


def _get_provider_config():
    """获取当前提供商的有效配置（含默认值）"""
    provider = api_config['provider']
    defaults = PROVIDER_DEFAULTS.get(provider, {})
    base_url = api_config['base_url'] or defaults.get('base_url', '')
    model = api_config['model'] or defaults.get('model', '')
    return provider, base_url, model


def _is_anthropic():
    return api_config['provider'] == 'anthropic'


@app.route('/api/config/test', methods=['POST'])
def test_config():
    """测试API连接（支持所有提供商）"""
    if not api_config['api_key'] or not api_config['provider']:
        return jsonify({'status': 'error', 'message': '请先配置API Key和提供商'})
    try:
        import requests
        provider, base_url, model = _get_provider_config()

        if _is_anthropic():
            resp = requests.post(
                f"{base_url}/v1/messages",
                headers={
                    "x-api-key": api_config['api_key'],
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json"
                },
                json={
                    "model": model,
                    "messages": [{"role": "user", "content": "回复ok两个字即可"}],
                    "max_tokens": 10
                },
                timeout=15
            )
            if resp.status_code == 200:
                reply = resp.json()['content'][0]['text']
                return jsonify({'status': 'ok', 'message': f'连接成功！模型回复: {reply}'})
            else:
                return jsonify({'status': 'error', 'message': f'请求失败 (HTTP {resp.status_code}): {resp.text[:200]}'})
        else:
            # 所有OpenAI兼容的提供商（deepseek/openai/qwen/zhipu/moonshot/baichuan/minimax/spark/ollama/custom）
            resp = requests.post(
                f"{base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_config['api_key']}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": model,
                    "messages": [{"role": "user", "content": "回复ok两个字即可"}],
                    "max_tokens": 10
                },
                timeout=15
            )
            if resp.status_code == 200:
                reply = resp.json()['choices'][0]['message']['content']
                return jsonify({'status': 'ok', 'message': f'连接成功！模型回复: {reply}'})
            else:
                return jsonify({'status': 'error', 'message': f'请求失败 (HTTP {resp.status_code}): {resp.text[:200]}'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'连接失败: {str(e)}'})

@app.route('/api/skills', methods=['GET'])
def list_skills():
    skills = []
    # 内置技能
    for d in SKILL_DIR.iterdir() if SKILL_DIR.exists() else []:
        if d.is_dir() and not d.name.startswith('_'):
            txt_path = d / 'skill.txt'
            desc = ''
            if txt_path.exists():
                desc = txt_path.read_text(encoding='utf-8')
            skills.append({
                'name': d.name,
                'type': 'builtin',
                'description': desc,
                'path': str(d)
            })
    # AI生成的技能
    for d in AI_SKILL_DIR.iterdir() if AI_SKILL_DIR.exists() else []:
        if d.is_dir() and not d.name.startswith('_'):
            txt_path = d / 'skill.txt'
            desc = ''
            if txt_path.exists():
                desc = txt_path.read_text(encoding='utf-8')
            skills.append({
                'name': d.name,
                'type': 'ai',
                'description': desc,
                'path': str(d)
            })
    return jsonify(skills)

@app.route('/api/skills/create', methods=['POST'])
def create_skill():
    # 支持JSON和FormData两种格式
    content_type = request.content_type or ''
    if 'multipart/form-data' in content_type:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        # 处理示例文件上传
        example_before = request.files.get('exampleBefore')
        example_after = request.files.get('exampleAfter')
        # 保存示例文件并读取数据
        example_before_data = ''
        example_after_data = ''
        if example_before:
            save_path = UPLOAD_DIR / ('example_before_' + example_before.filename)
            example_before.save(str(save_path))
            example_before_data = read_excel_summary(str(save_path))
        if example_after:
            save_path = UPLOAD_DIR / ('example_after_' + example_after.filename)
            example_after.save(str(save_path))
            example_after_data = read_excel_summary(str(save_path))
    else:
        data = request.json or {}
        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        example_before_data = ''
        example_after_data = ''

    if not name or not description:
        return jsonify({'error': '技能名称和描述不能为空'}), 400
    path, error = try_generate_skill_with_ai(name, description, example_before_data, example_after_data)
    if error:
        return jsonify({'error': error}), 400
    return jsonify({'status': 'ok', 'name': name, 'path': path, 'message': f'技能 "{name}" 创建成功'})


def read_excel_summary(file_path):
    """读取Excel/TXT文件的摘要信息（表头+前3行数据），供AI参考"""
    try:
        ext = os.path.splitext(file_path)[1].lower()
        rows = []
        if ext == '.txt':
            # 读取txt文件，自动检测编码和分隔符
            encodings = ['utf-8', 'gbk', 'gb2312', 'latin-1']
            lines = []
            for enc in encodings:
                try:
                    with open(file_path, 'r', encoding=enc) as f:
                        lines = f.readlines()
                    if lines:
                        break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            if not lines:
                return '无法读取txt文件'
            # 检测分隔符
            delimiter = '\t'
            for delim in ['\t', ',', ' ', '|', ';']:
                if delim in lines[0]:
                    delimiter = delim
                    break
            for line in lines[:4]:
                parts = [p.strip()[:50] for p in line.strip().split(delimiter) if p.strip()]
                if parts:
                    rows.append(parts)
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            for r in range(min(ws.nrows, 4)):
                row = [str(ws.cell(r, c).value)[:50] for c in range(ws.ncols)]
                rows.append(row)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 4:
                    break
                rows.append([str(v)[:50] if v is not None else '' for v in row])
        if not rows:
            return ''
        header = rows[0]
        lines = ['表头: ' + ' | '.join(header)]
        for i, row in enumerate(rows[1:], 1):
            lines.append(f'行{i}: ' + ' | '.join(row))
        return '\n'.join(lines)
    except Exception as e:
        return f'读取失败: {e}'

@app.route('/api/process', methods=['POST'])
def process_files():
    skill_name = request.form.get('skill', '').strip()
    if not skill_name:
        return jsonify({'error': '未选择技能'}), 400

    # 查找技能处理函数
    process_func = None
    skill_path = SKILL_DIR / skill_name / 'handler.py'
    ai_skill_path = AI_SKILL_DIR / skill_name / 'handler.py'

    if skill_path.exists():
        module_name = f"skill_{skill_name}"
        spec = importlib.util.spec_from_file_location(module_name, str(skill_path))
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        process_func = module.process
    elif ai_skill_path.exists():
        module_name = f"ai_skill_{skill_name}"
        spec = importlib.util.spec_from_file_location(module_name, str(ai_skill_path))
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        process_func = module.process
    else:
        return jsonify({'error': f'未找到技能: {skill_name}'}), 400

    # 保存上传的文件
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': '未上传文件'}), 400

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results = []

    for f in files:
        if not f.filename:
            continue
        # 保存到临时目录
        save_name = f.filename
        save_path = UPLOAD_DIR / save_name
        f.save(str(save_path))

        # 如果是txt文件，检测编码并转换为UTF-8，确保技能代码能正确读取
        if save_path.suffix.lower() == '.txt':
            try:
                encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'big5', 'latin-1']
                content = None
                for enc in encodings:
                    try:
                        with open(save_path, 'r', encoding=enc) as f_txt:
                            content = f_txt.read()
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                if content is not None:
                    with open(save_path, 'w', encoding='utf-8') as f_txt:
                        f_txt.write(content)
            except Exception:
                pass  # 编码转换失败，保持原文件

        # 构建输出路径：原名_时间戳_after.xlsx
        stem = os.path.splitext(save_name)[0]
        output_name = f"{stem}_{timestamp}_after.xlsx"
        output_path = OUTPUT_DIR / output_name

        try:
            # 确保输出目录存在
            OUTPUT_DIR.mkdir(exist_ok=True)
            try:
                result_path, image_paths = process_func(str(save_path), str(output_path))
            except ImportError as ie:
                missing_mod = str(ie).split("'")[1] if "'" in str(ie) else str(ie)
                results.append({
                    'original': save_name,
                    'status': 'error',
                    'message': f'缺少依赖库 {missing_mod}，请运行: pip install {missing_mod}'
                })
                continue
            if result_path and os.path.exists(result_path):
                # 优先使用技能函数返回的图片路径
                chart_images = []
                if image_paths:
                    for img in image_paths:
                        if os.path.exists(img):
                            chart_images.append(img)
                # 如果技能函数未返回图片，尝试搜索 output 目录
                if not chart_images:
                    output_dir = os.path.dirname(result_path)
                    import glob
                    # 优先搜索与Excel同名的PNG图片
                    excel_stem = os.path.splitext(os.path.basename(result_path))[0]
                    search_patterns = [
                        f"{excel_stem}.png",
                        f"{stem}*chart*", f"{stem}*图*", f"{stem}*image*", f"{stem}*plot*",
                        "*chart*", "*图*", "*plot*",
                    ]
                    for pattern in search_patterns:
                        for img in glob.glob(os.path.join(output_dir, pattern)):
                            if img not in chart_images and not img.endswith('.xlsx') and not img.endswith('.xls'):
                                chart_images.append(img)
                    for img in glob.glob(os.path.join(output_dir, "*.png")):
                        if img not in chart_images:
                            chart_images.append(img)
                    for img in glob.glob(os.path.join(output_dir, "*.jpg")):
                        if img not in chart_images:
                            chart_images.append(img)

                results.append({
                    'original': save_name,
                    'output': output_name,
                    'output_url': f'/api/download/{output_name}',
                    'images': [
                        {'url': f'/api/image/{os.path.basename(img)}',
                         'name': os.path.basename(img)}
                        for img in chart_images
                    ],
                    'status': 'success'
                })
            else:
                results.append({
                    'original': save_name,
                    'status': 'error',
                    'message': '处理失败，未生成输出文件'
                })
        except Exception as e:
            results.append({
                'original': save_name,
                'status': 'error',
                'message': str(e)
            })

    return jsonify({'results': results})

@app.route('/api/download/<filename>')
def download_file(filename):
    """下载文件（强制下载）"""
    # 先从output目录查找
    file_path = OUTPUT_DIR / filename
    if file_path.exists():
        return send_from_directory(str(OUTPUT_DIR), filename, as_attachment=True)
    # 再从uploads目录查找
    file_path = UPLOAD_DIR / filename
    if file_path.exists():
        return send_from_directory(str(UPLOAD_DIR), filename, as_attachment=True)
    return jsonify({'error': '文件不存在'}), 404

@app.route('/api/image/<filename>')
def view_image(filename):
    """查看图片（直接显示）"""
    # 先从output目录查找
    file_path = OUTPUT_DIR / filename
    if file_path.exists():
        return send_from_directory(str(OUTPUT_DIR), filename)
    return jsonify({'error': '文件不存在'}), 404

@app.route('/api/skills/code/<skill_name>', methods=['GET'])
def get_skill_code(skill_name):
    """获取技能的handler.py代码，用于修改"""
    ai_path = AI_SKILL_DIR / skill_name / 'handler.py'
    skill_path = SKILL_DIR / skill_name / 'handler.py'
    if ai_path.exists():
        code = ai_path.read_text(encoding='utf-8')
        return jsonify({'name': skill_name, 'code': code, 'type': 'ai'})
    elif skill_path.exists():
        code = skill_path.read_text(encoding='utf-8')
        return jsonify({'name': skill_name, 'code': code, 'type': 'builtin'})
    return jsonify({'error': f'未找到技能: {skill_name}'}), 404


@app.route('/api/skills/modify', methods=['POST'])
def modify_skill():
    """通过多轮对话修改已有技能"""
    data = request.json or {}
    skill_name = data.get('name', '').strip()
    instruction = data.get('instruction', '').strip()
    current_code = data.get('current_code', '')

    if not skill_name or not instruction:
        return jsonify({'error': '技能名称和修改说明不能为空'}), 400

    if not api_config['api_key'] or not api_config['provider']:
        return jsonify({'error': '未配置API Key，请先配置'}), 400

    system_prompt = """你是一个Python代码修改专家。用户会给你一段现有的Excel处理技能代码和修改要求，你需要根据修改要求更新代码。

要求：
1. 在现有代码基础上进行修改，保留不需要改动的部分
2. 函数签名必须保持为 process(input_path, output_path)
3. 返回 (output_path, image_paths_list)
4. 使用 openpyxl 处理 .xlsx，xlrd 处理 .xls
5. 必须支持 .xls、.xlsx 和 .txt 三种格式
6. 日志使用 logging 模块
7. import 语句放在函数内部
8. 只返回纯Python代码，不要包含任何markdown标记、注释或解释
9. 如果需要生成图片，使用 matplotlib，保存为PNG
"""

    user_prompt = f"""现有技能代码：
```python
{current_code}
```

修改要求：{instruction}

请返回修改后的完整Python代码，不要任何解释。"""

    try:
        import requests as _requests
        provider, base_url, model = _get_provider_config()

        if _is_anthropic():
            resp = _requests.post(
                f"{base_url}/v1/messages",
                headers={
                    "x-api-key": api_config['api_key'],
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json"
                },
                json={
                    "model": model,
                    "system": system_prompt,
                    "messages": [{"role": "user", "content": user_prompt}],
                    "max_tokens": 4096,
                    "temperature": 0.1
                },
                timeout=60
            )
            resp.raise_for_status()
            new_code = resp.json()['content'][0]['text']
        else:
            resp = _requests.post(
                f"{base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_config['api_key']}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": model,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt}
                    ],
                    "temperature": 0.1,
                    "max_tokens": 4096
                },
                timeout=60
            )
            resp.raise_for_status()
            new_code = resp.json()['choices'][0]['message']['content']
    except Exception as e:
        return jsonify({'error': f'AI调用失败: {str(e)}'}), 400

    if not new_code:
        return jsonify({'error': 'AI未返回代码'}), 400

    # 清理代码块标记
    new_code = new_code.strip()
    if new_code.startswith('```'):
        lines = new_code.split('\n')
        start = 1 if lines[0].startswith('```') else 0
        end = len(lines) - 1 if lines[-1].strip() == '```' else len(lines)
        new_code = '\n'.join(lines[start:end])

    # 验证语法
    try:
        compile(new_code, '<string>', 'exec')
    except SyntaxError as e:
        return jsonify({'error': f'生成的代码有语法错误: {e}'}), 400

    # 保存更新后的代码
    skill_path = AI_SKILL_DIR / skill_name
    skill_path.mkdir(exist_ok=True)
    with open(skill_path / 'handler.py', 'w', encoding='utf-8') as f:
        f.write(new_code)

    # 更新skill.txt中的修改记录
    txt_path = skill_path / 'skill.txt'
    txt_content = ''
    if txt_path.exists():
        txt_content = txt_path.read_text(encoding='utf-8')
    txt_content += f"\n\n修改记录 [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]: {instruction}"
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write(txt_content)

    return jsonify({
        'status': 'ok',
        'name': skill_name,
        'code': new_code,
        'message': f'技能 "{skill_name}" 修改成功'
    })


@app.route('/api/skills/template', methods=['GET'])
def get_skill_template():
    """下载技能开发规范模板"""
    # 模板文件从只读资源目录读取
    template_path = STATIC_SRC_DIR / 'ai_skills' / '技能开发规范.txt'
    if not template_path.exists():
        template_path = AI_SKILL_DIR / '技能开发规范.txt'
    if template_path.exists():
        return send_file(str(template_path), as_attachment=True,
                         download_name='技能开发规范.txt',
                         mimetype='text/plain')
    return jsonify({'error': '模板文件不存在'}), 404


@app.route('/api/pending/<skill_name>')
def list_pending(skill_name):
    skill_dir = PENDING_DIR / skill_name
    if not skill_dir.exists():
        return jsonify([])
    files = []
    for f in skill_dir.iterdir():
        if f.is_file() and not f.name.startswith('~$') and not f.name.startswith('.'):
            files.append({
                'name': f.name,
                'size': f.stat().st_size,
                'url': f'/api/pending-file/{skill_name}/{f.name}'
            })
    return jsonify(files)

@app.route('/api/pending-file/<skill_name>/<filename>')
def download_pending(skill_name, filename):
    skill_dir = PENDING_DIR / skill_name
    file_path = skill_dir / filename
    if not file_path.exists():
        return jsonify({'error': '文件不存在'}), 404
    return send_from_directory(str(skill_dir), filename, as_attachment=True)

@app.route('/api/upload-to-pending/<skill_name>', methods=['POST'])
def upload_to_pending(skill_name):
    skill_dir = PENDING_DIR / skill_name
    skill_dir.mkdir(parents=True, exist_ok=True)
    files = request.files.getlist('files')
    uploaded = []
    for f in files:
        if f.filename:
            f.save(str(skill_dir / f.filename))
            uploaded.append(f.filename)
    return jsonify({'uploaded': uploaded})

@app.route('/api/chat', methods=['POST'])
def chat_with_ai():
    """与AI对话聊天"""
    data = request.json or {}
    message = data.get('message', '').strip()
    history = data.get('history', [])
    if not message:
        return jsonify({'error': '消息不能为空'}), 400
    if not api_config['api_key'] or not api_config['provider']:
        return jsonify({'error': '未配置API Key，请先在上方配置'}), 400

    system_prompt = "你是一个友好的AI助手，擅长Excel数据处理、Python编程、数据分析等领域。请用中文回复，回答简洁明了。"

    messages = [{"role": "system", "content": system_prompt}]
    # 加入历史对话
    for h in history[-10:]:  # 最多保留最近10轮
        messages.append({"role": h.get('role', 'user'), "content": h.get('content', '')})
    messages.append({"role": "user", "content": message})

    try:
        import requests as _requests
        provider, base_url, model = _get_provider_config()

        if _is_anthropic():
            # Anthropic格式需要把system单独传
            system_msg = messages[0]['content']
            chat_msgs = [m for m in messages[1:]]
            resp = _requests.post(
                f"{base_url}/v1/messages",
                headers={
                    "x-api-key": api_config['api_key'],
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json"
                },
                json={
                    "model": model,
                    "system": system_msg,
                    "messages": chat_msgs,
                    "max_tokens": 2048,
                    "temperature": 0.7
                },
                timeout=30
            )
            resp.raise_for_status()
            reply = resp.json()['content'][0]['text']
        else:
            resp = _requests.post(
                f"{base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_config['api_key']}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": model,
                    "messages": messages,
                    "temperature": 0.7,
                    "max_tokens": 2048
                },
                timeout=30
            )
            resp.raise_for_status()
            reply = resp.json()['choices'][0]['message']['content']

        return jsonify({'reply': reply})
    except Exception as e:
        return jsonify({'error': f'AI调用失败: {str(e)}'}), 400


@app.route('/api/skills/rename', methods=['POST'])
def rename_skill():
    """重命名技能"""
    data = request.json or {}
    old_name = data.get('old_name', '').strip()
    new_name = data.get('new_name', '').strip()
    if not old_name or not new_name:
        return jsonify({'error': '旧名称和新名称不能为空'}), 400
    if old_name == new_name:
        return jsonify({'error': '新名称与旧名称相同'}), 400

    # 检查新名称是否已存在
    if (AI_SKILL_DIR / new_name).exists() or (SKILL_DIR / new_name).exists():
        return jsonify({'error': f'技能 "{new_name}" 已存在'}), 400

    # 查找技能目录（优先ai_skills）
    old_ai_path = AI_SKILL_DIR / old_name
    old_builtin_path = SKILL_DIR / old_name
    if old_ai_path.exists() and old_ai_path.is_dir():
        old_ai_path.rename(AI_SKILL_DIR / new_name)
    elif old_builtin_path.exists() and old_builtin_path.is_dir():
        old_builtin_path.rename(SKILL_DIR / new_name)
    else:
        return jsonify({'error': f'未找到技能: {old_name}'}), 404

    return jsonify({'status': 'ok', 'message': f'技能 "{old_name}" 已重命名为 "{new_name}"'})


@app.route('/api/skills/delete', methods=['POST'])
def delete_skill():
    """删除技能"""
    data = request.json or {}
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': '技能名称不能为空'}), 400

    ai_path = AI_SKILL_DIR / name
    builtin_path = SKILL_DIR / name

    if ai_path.exists() and ai_path.is_dir():
        shutil.rmtree(ai_path)
        return jsonify({'status': 'ok', 'message': f'技能 "{name}" 已删除'})
    elif builtin_path.exists() and builtin_path.is_dir():
        shutil.rmtree(builtin_path)
        return jsonify({'status': 'ok', 'message': f'技能 "{name}" 已删除'})
    else:
        return jsonify({'error': f'未找到技能: {name}'}), 404


if __name__ == '__main__':
    import threading
    print("=" * 50)
    print("  Excel Agent Online")
    print(f"  访问地址: http://localhost:5000")
    print(f"  技能目录: {SKILL_DIR}")
    print(f"  待处理目录: {PENDING_DIR}")
    print(f"  输出目录: {OUTPUT_DIR}")
    print("=" * 50)
    # 自动打开浏览器
    def open_browser():
        import time as _time
        _time.sleep(1.5)
        webbrowser.open('http://localhost:5000')
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='0.0.0.0', port=5000, debug=False)
