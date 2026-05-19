"""
温升技能 - 处理温升相关Excel文件

功能:
1. 将文本类型的数字转换为数值类型
2. 统一时间列格式（第一行保留完整日期时间，后续行只保留时间部分）
3. 对数值列插入折线图（横坐标为时间列）

支持的文件格式:
- .xls（旧版Excel）
- .xlsx（新版Excel）

支持的列结构:
- 必须有"时间"列（表头包含"时间"关键字）
- 其余数值列自动生成折线图
- 非数值文本列（如编号、热电偶类型等）自动跳过
"""

import os
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from datetime import datetime


def read_excel_data(input_path):
    """
    读取Excel文件，支持 .xls 和 .xlsx 格式。
    返回 (sheet_name, all_data, is_xls)
    all_data 是二维列表，第一行为表头。
    """
    ext = os.path.splitext(input_path)[1].lower()

    if ext == '.xls':
        wb_in = xlrd.open_workbook(input_path)
        sheet_in = wb_in.sheet_by_index(0)
        sheet_name = sheet_in.name
        nrows = sheet_in.nrows
        ncols = sheet_in.ncols
        all_data = []
        for r in range(nrows):
            row = []
            for c in range(ncols):
                cell = sheet_in.cell(r, c)
                row.append(cell.value)
            all_data.append(row)
        return sheet_name, all_data, True

    else:
        # .xlsx 格式
        wb_in = load_workbook(input_path, data_only=True)
        ws_in = wb_in.active
        sheet_name = ws_in.title
        all_data = []
        for row in ws_in.iter_rows(values_only=True):
            all_data.append(list(row))
        return sheet_name, all_data, False


def parse_datetime_value(value):
    """
    解析各种格式的日期时间值。
    支持：
    - Excel float 序列号（xlrd 读取的日期时间）
    - 字符串格式：2025-07-16 14:53:27.403, 22:07:31.623, 2025-7-16 14:53 等
    - datetime 对象
    返回 (datetime_obj, is_valid)
    """
    if value is None:
        return None, False

    # 如果已经是 datetime 对象
    if isinstance(value, datetime):
        return value, True

    # xlrd float 日期序列号
    if isinstance(value, float):
        # 小于 100000 的浮点数可能是什么？
        # Excel 日期序列号从 1（1900-01-01）开始，现代日期大约在 40000-50000 范围
        # 但如果只有时间，值会很小（0-1之间）
        # 我们不把它当日期处理，直接返回
        return None, False

    if isinstance(value, int):
        # 整数可能是编号（1, 2, 3...），不作为日期
        return None, False

    if isinstance(value, str):
        value = value.strip()
        if value == '' or value == '--' or value == '-':
            return None, False

        # 尝试各种日期时间格式
        formats = [
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%H:%M:%S.%f",
            "%H:%M:%S",
            "%H:%M",
            "%Y/%m/%d %H:%M:%S.%f",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt), True
            except ValueError:
                continue

        # 尝试处理 "2025-7-16" 这种月/日无前导零的格式
        # 替换为标准格式再试
        import re
        # 匹配类似 "22:07:31.623" 的纯时间（已有格式覆盖，这里处理混合格式）
        # 匹配 "2025-7-16 14:53:27" 格式
        m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})\s+(\d{1,2}):(\d{1,2}):(\d{1,2})(\.(\d+))?', value)
        if m:
            year = int(m.group(1))
            month = int(m.group(2))
            day = int(m.group(3))
            hour = int(m.group(4))
            minute = int(m.group(5))
            second = int(m.group(6))
            microsecond = int(m.group(8)) if m.group(8) else 0
            if m.group(8):
                microsecond = int(m.group(8).ljust(6, '0')[:6])
            try:
                return datetime(year, month, day, hour, minute, second, microsecond), True
            except ValueError:
                pass

        return None, False

    return None, False


def is_time_column(header):
    """判断一列是否是时间列（表头包含'时间'关键字）"""
    header_str = str(header).strip()
    return '时间' in header_str


def is_numeric_column(all_data, col_index, header):
    """
    判断一列是否为数值列。
    规则：排除编号列和文本类型列，检查数据行中是否有足够的数值。
    """
    header_str = str(header).strip()

    # 排除明显的非数值列
    if header_str == '编号' or header_str == '序号' or '热电偶类型' in header_str:
        return False

    # 如果表头看起来像数值列（包含℃、温度、℃等关键字），直接判断为数值列
    if any(kw in header_str for kw in ['℃', '°C', '温度', '温度', 'T1', 'T2', 'T3', 'T4',
                                        'T5', 'T6', 'T7', 'T8', 'T9', 'T10',
                                        '主芯片', 'WIFI', '环温', '芯片', '板']):
        return True

    # 检查数据行中的值，如果大部分是数值则判定为数值列
    numeric_count = 0
    total_count = 0
    for r in range(1, min(len(all_data), 20)):  # 检查前20行数据
        if r >= len(all_data) or col_index >= len(all_data[r]):
            continue
        val = all_data[r][col_index]
        if val is None or (isinstance(val, str) and val.strip() in ['', '--', '-']):
            continue
        total_count += 1
        if isinstance(val, (int, float)):
            numeric_count += 1
        elif isinstance(val, str):
            try:
                float(val.strip())
                numeric_count += 1
            except ValueError:
                pass

    # 如果 70% 以上的非空值是数值，则认为是数值列
    if total_count > 0 and numeric_count / total_count >= 0.7:
        return True

    return False


def format_datetime_for_first_row(dt):
    """格式化第一行时间：
    如果有日期部分：2025/7/16 14:53:27
    如果只有时间部分：22:07:31.623
    """
    # 判断是否只有时间部分（年月日都是默认值）
    # 我们检查原始数据中是否有日期部分
    # 简化处理：如果 datetime 的 date 是 1900-01-01，认为只有时间
    if dt.year == 1900 and dt.month == 1 and dt.day == 1:
        # 只有时间部分
        if dt.microsecond > 0:
            return f"{dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}.{dt.microsecond // 1000:03d}"
        else:
            return f"{dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}"
    else:
        # 有完整日期
        return f"{dt.year}/{dt.month}/{dt.day} {dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}"


def format_time_only(dt):
    """格式化后续行时间：14:53:27 或 22:07:31.623"""
    if dt.microsecond > 0:
        return f"{dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}.{dt.microsecond // 1000:03d}"
    else:
        return f"{dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}"


def try_convert_number(value):
    """尝试将值转换为数字"""
    if isinstance(value, (int, float)):
        return value, True

    if isinstance(value, str):
        value = value.strip()
        if value == '' or value == '--' or value == '-' or value == 'N/A':
            return value, False
        try:
            return float(value), True
        except ValueError:
            return value, False

    return value, False


def find_header_columns(headers, all_data):
    """
    分析表头，智能识别各列类型。
    - 时间列：表头包含"时间"
    - 数值列：包含温度关键字或数据大部分为数值
    - 其他列：编号、热电偶类型等文本列
    """
    result = {
        'time_col': -1,
        'value_cols': [],
        'skip_cols': [],  # 编号、热电偶类型等跳过的列
    }

    # 先找时间列
    for i, header in enumerate(headers):
        if is_time_column(header):
            result['time_col'] = i
            break

    # 找数值列和跳过列
    for i, header in enumerate(headers):
        if i == result['time_col']:
            continue  # 时间列单独处理
        if is_numeric_column(all_data, i, header):
            result['value_cols'].append(i)
        else:
            result['skip_cols'].append(i)

    return result


def process(input_path, output_path):
    """处理Excel文件 - 温升技能（通用版）"""
    # 读取Excel数据
    sheet_name, all_data, is_xls = read_excel_data(input_path)

    nrows = len(all_data)
    if nrows == 0:
        print("    文件为空，跳过处理")
        return None

    ncols = len(all_data[0]) if nrows > 0 else 0

    # 分析表头
    headers = all_data[0]
    col_info = find_header_columns(headers, all_data)

    if col_info['time_col'] < 0:
        print("    未找到时间列，跳过处理")
        return None

    if not col_info['value_cols']:
        print("    未找到数值列，跳过处理")
        return None

    print(f"    时间列: 第{col_info['time_col'] + 1}列 ({headers[col_info['time_col']]})")
    print(f"    数值列: {[str(headers[i]) for i in col_info['value_cols']] if col_info['value_cols'] else '无'}")

    # 创建输出工作簿
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = sheet_name

    # 写入表头
    for c, header in enumerate(headers):
        ws_out.cell(row=1, column=c + 1, value=header)

    # 处理数据行
    value_col_indices = []

    for r in range(1, nrows):
        row = all_data[r]
        for c in range(ncols):
            if c >= len(row):
                continue
            value = row[c]

            # 时间列处理
            if c == col_info['time_col']:
                dt, is_valid = parse_datetime_value(value)
                if is_valid and dt:
                    if r == 1:
                        # 第一行：格式化时间（完整日期时间 或 时间）
                        ws_out.cell(row=r + 1, column=c + 1,
                                    value=format_datetime_for_first_row(dt))
                    else:
                        # 后续行：只保留时间部分
                        ws_out.cell(row=r + 1, column=c + 1,
                                    value=format_time_only(dt))
                else:
                    # 无法解析为时间，保持原样
                    ws_out.cell(row=r + 1, column=c + 1, value=value)

            # 数值列处理
            elif c in col_info['value_cols']:
                converted, is_number = try_convert_number(value)
                ws_out.cell(row=r + 1, column=c + 1, value=converted)
                # 记录第二行（第一个数据行）中的数值列
                if r == 1 and c not in value_col_indices:
                    value_col_indices.append(c)

            # 其他列（编号、热电偶类型等）保持原样
            else:
                ws_out.cell(row=r + 1, column=c + 1, value=value)

    # 插入折线图（如果有数值列）
    if value_col_indices and col_info['time_col'] >= 0:
        chart = LineChart()
        chart.title = "温升温度曲线"
        chart.y_axis.title = "温度 (℃)"
        chart.x_axis.title = "时间"
        chart.style = 10
        chart.width = 25
        chart.height = 15

        # X轴数据（时间列）
        x_data = Reference(ws_out,
                           min_col=col_info['time_col'] + 1,
                           min_row=2,
                           max_row=nrows)

        # 添加数值列作为数据系列
        for vc in value_col_indices:
            data = Reference(ws_out,
                             min_col=vc + 1,
                             min_row=1,
                             max_row=nrows)
            chart.add_data(data, titles_from_data=True)

        chart.set_categories(x_data)

        # 设置图表位置（放在数据下方）
        ws_out.add_chart(chart, f"A{nrows + 3}")

    # 保存输出文件（确保输出为xlsx格式）
    if not output_path.endswith('.xlsx'):
        output_path = output_path.rsplit('.', 1)[0] + '.xlsx'
    wb_out.save(output_path)
    print(f"    处理完成: {output_path}")

    # 生成matplotlib独立PNG图片（截图形式展示在网页图片区）
    image_paths = []
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        # 设置中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False

        fig, ax = plt.subplots(figsize=(16, 8))
        colors = ['#e74c3c', '#3498db', '#2ecc71', '#f39c12', '#9b59b6', '#1abc9c', '#e67e22', '#34495e']

        time_labels = []
        for r in range(2, nrows + 1):
            val = ws_out.cell(row=r, column=col_info['time_col'] + 1).value
            time_labels.append(str(val) if val else '')

        # 每隔N个点显示一个标签，避免拥挤
        max_labels = 30
        step = max(1, len(time_labels) // max_labels)

        # 横坐标使用时间列
        x_values = list(range(len(time_labels)))
        
        for idx, vc in enumerate(value_col_indices):
            header_name = str(headers[vc])
            y_values = []
            for r in range(2, nrows + 1):
                val = ws_out.cell(row=r, column=vc + 1).value
                if isinstance(val, (int, float)):
                    y_values.append(float(val))
                elif isinstance(val, str):
                    try:
                        y_values.append(float(val.strip()))
                    except ValueError:
                        y_values.append(None)
                else:
                    y_values.append(None)

            color = colors[idx % len(colors)]
            ax.plot(x_values, y_values, '-', color=color, linewidth=1.5, label=header_name, markersize=1.5)

        ax.set_title('温升温度曲线', fontsize=18, fontweight='bold')
        ax.set_xlabel('时间', fontsize=12)
        ax.set_ylabel('温度 (℃)', fontsize=12)
        ax.legend(fontsize=10, loc='best', framealpha=0.9)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(0, max(1, len(time_labels) - 1))
        
        # 横坐标刻度显示实际时间值（最多显示15个标签避免拥挤）
        num_points = len(time_labels)
        tick_count = min(15, num_points)
        if num_points > 1:
            tick_indices = [int(i * (num_points - 1) / (tick_count - 1)) for i in range(tick_count)]
        else:
            tick_indices = [0]
        ax.set_xticks(tick_indices)
        ax.set_xticklabels([time_labels[i] for i in tick_indices], rotation=45, ha='right', fontsize=8)

        plt.tight_layout()

        output_dir = os.path.dirname(output_path)
        # 图片名与Excel输出文件名保持一致（扩展名改为.png）
        base_name = os.path.splitext(os.path.basename(output_path))[0]
        png_path = os.path.join(output_dir, f'{base_name}.png')
        fig.savefig(png_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        image_paths.append(png_path)
        print(f"    PNG图表已保存: {png_path}")
    except Exception as e:
        import logging
        logging.warning(f"matplotlib生成PNG图表失败: {e}")

    return output_path, image_paths
