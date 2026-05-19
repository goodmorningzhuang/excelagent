import os

def process(input_path, output_path):
    import os
    import logging
    from datetime import datetime

    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

    input_ext = os.path.splitext(input_path)[1].lower()
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)

    # 生成带时间戳的文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = os.path.splitext(os.path.basename(output_path))[0]
    final_output_name = f"{base_name}_{timestamp}.xlsx"
    final_output_path = os.path.join(output_dir, final_output_name)

    from openpyxl import Workbook, load_workbook

    # ========== 第一步：将数据写入Excel ==========
    if input_ext == '.txt':
        # === TXT文件：读取内容，分列，写入Excel ===
        wb = Workbook()
        ws = wb.active
        ws.title = "数据"

        # 多编码读取txt
        lines = None
        for enc in ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin-1']:
            try:
                with open(input_path, 'r', encoding=enc) as f:
                    lines = f.readlines()
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        if lines is None:
            with open(input_path, 'r', encoding='utf-8', errors='replace') as f:
                lines = f.readlines()

        if not lines:
            logging.error("txt文件为空")
            return None, []

        import re

        # 逐行分列写入Excel（用正则按空白分列，兼容多空格/Tab混合）
        for line in lines:
            line = line.strip()
            if not line:
                continue
            parts = re.split(r'\s+', line)
            row = [p.strip() for p in parts]
            ws.append(row)

        logging.info(f"txt文件已写入Excel，共{ws.max_row}行")

        # 将文本类型的数字转换为真正的数字
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    val = cell.value.strip()
                    # 去掉百分号
                    if val.endswith('%'):
                        try:
                            cell.value = float(val[:-1])
                            continue
                        except ValueError:
                            pass
                    # 尝试转为数字
                    try:
                        cell.value = int(val)
                        continue
                    except ValueError:
                        pass
                    try:
                        cell.value = float(val)
                    except ValueError:
                        pass

    else:
        # === Excel文件（.xls或.xlsx）：直接读取 ===
        if input_ext == '.xls':
            import xlrd
            workbook = xlrd.open_workbook(input_path)
            sheet = workbook.sheet_by_index(0)
            wb = Workbook()
            ws = wb.active
            ws.title = sheet.name
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        row_data.append(datetime(*xlrd.xldate_as_tuple(cell.value, workbook.datemode)))
                    else:
                        row_data.append(cell.value)
                ws.append(row_data)
        else:
            wb = load_workbook(input_path)
            ws = wb.active

        logging.info(f"Excel文件已读取，共{ws.max_row}行")

    # ========== 第二步：识别时间/电流/电压/电量列 ==========
    header_row = [cell.value for cell in ws[1]]
    logging.info(f"表头: {header_row}")

    def find_col(keywords):
        for i, h in enumerate(header_row):
            if h:
                h_str = str(h).strip()
                for kw in keywords:
                    if kw in h_str:
                        return i
        return None

    time_col = find_col(['时间', 'time', 'Time', 'TIME'])
    current_col = find_col(['电流', 'current', 'Current', 'CURRENT'])
    voltage_col = find_col(['电压', 'voltage', 'Voltage', 'VOLTAGE'])
    capacity_col = find_col(['电量', 'capacity', 'Capacity', 'CAPACITY', 'SOC', 'soc'])

    logging.info(f"时间列: {time_col}, 电流列: {current_col}, 电压列: {voltage_col}, 电量列: {capacity_col}")

    if time_col is None or current_col is None or voltage_col is None or capacity_col is None:
        logging.warning(f"未找到完整的四列（时间/电流/电压/电量），当前找到: 时间={time_col}, 电流={current_col}, 电压={voltage_col}, 电量={capacity_col}")
        wb.save(final_output_path)
        return final_output_path, []

    # ========== 第三步：提取数据并制作折线图 ==========
    times = []
    currents = []
    voltages = []
    capacities = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) <= max(time_col, current_col, voltage_col, capacity_col):
            continue
        try:
            t = str(row[time_col]) if row[time_col] is not None else ''
            c = float(str(row[current_col]).replace('%', '').strip())
            v = float(str(row[voltage_col]).replace('%', '').strip())
            cap = float(str(row[capacity_col]).replace('%', '').strip())
            if t.strip():
                times.append(t)
                currents.append(c)
                voltages.append(v)
                capacities.append(cap)
        except (ValueError, TypeError, IndexError):
            continue

    logging.info(f"提取数据行数: {len(times)}")

    image_paths = []
    if len(times) < 2:
        logging.warning("数据行数不足，无法生成图表")
        wb.save(final_output_path)
        return final_output_path, []

    # 生成折线图
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    fig, ax = plt.subplots(figsize=(14, 6))

    x = list(range(len(times)))

    # 三合一折线图
    ax.plot(x, currents, 'b-', linewidth=1.2, label='电流 (A)')
    ax.plot(x, voltages, 'r-', linewidth=1.2, label='电压 (V)')
    ax.plot(x, capacities, 'g-', linewidth=1.2, label='电量 (%)')

    ax.set_xlabel('时间', fontsize=11)
    ax.set_ylabel('数值', fontsize=11)
    ax.set_title('电流 / 电压 / 电量 vs 时间', fontsize=13, fontweight='bold')
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)

    # 横坐标显示实际时间（最多15个标签避免拥挤）
    num_points = len(times)
    tick_count = min(15, num_points)
    if num_points > 1:
        tick_indices = [int(i * (num_points - 1) / (tick_count - 1)) for i in range(tick_count)]
    else:
        tick_indices = [0]
    ax.set_xticks(tick_indices)
    ax.set_xticklabels([times[i] for i in tick_indices], rotation=45, ha='right', fontsize=8)

    plt.tight_layout()

    # 图片名与Excel输出文件名保持一致
    base_name_img = os.path.splitext(os.path.basename(final_output_path))[0]
    image_path = os.path.join(output_dir, f'{base_name_img}.png')
    fig.savefig(image_path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    image_paths.append(image_path)
    logging.info(f"PNG图表已保存: {image_path}")

    # ========== 第四步：在Excel中插入三合一折线图 ==========
    from openpyxl.chart import LineChart, Reference

    nrows = ws.max_row

    # 三合一折线图（横坐标时间，纵坐标电流+电压+电量）
    chart = LineChart()
    chart.title = "电流 / 电压 / 电量 vs 时间"
    chart.y_axis.title = "数值"
    chart.x_axis.title = "时间"
    chart.style = 10
    chart.width = 30
    chart.height = 18

    # X轴数据（时间列）
    x_data = Reference(ws, min_col=time_col + 1, min_row=2, max_row=nrows)

    # 添加三组数据（电流、电压、电量）
    current_data = Reference(ws, min_col=current_col + 1, min_row=1, max_row=nrows)
    voltage_data = Reference(ws, min_col=voltage_col + 1, min_row=1, max_row=nrows)
    capacity_data = Reference(ws, min_col=capacity_col + 1, min_row=1, max_row=nrows)

    chart.add_data(current_data, titles_from_data=True)
    chart.add_data(voltage_data, titles_from_data=True)
    chart.add_data(capacity_data, titles_from_data=True)
    chart.set_categories(x_data)

    # 设置图表位置（放在数据下方）
    ws.add_chart(chart, f"A{nrows + 3}")

    # 保存Excel（图表已嵌入）
    wb.save(final_output_path)
    logging.info(f"Excel文件已保存（含三合一折线图）: {final_output_path}")

    return final_output_path, image_paths
